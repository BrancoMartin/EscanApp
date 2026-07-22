"""Importacion masiva de productos desde Excel, CSV y exports de otros POS.

Spec: `openspec/specs/product import/spec.md`.

El objetivo es que un comerciante que ya tiene su lista de productos en otro
sistema (una planilla de Excel, un CSV, un export de Maxirest/Tango u otro POS)
pueda cargarla de una sola vez, en vez de tipear producto por producto.

Cada formato termina siendo, en el fondo, una tabla con encabezados. En lugar
de escribir un parser por marca de POS, resolvemos las columnas por ALIAS: se
mira el encabezado y se lo mapea al campo interno (barcode, nombre, precio,
descripcion, proveedor). Asi un mismo lector cubre Excel, CSV y la mayoria de
los exports, y sumar un POS nuevo es agregar un alias, no un parser.

La validacion de cada fila se delega en `ProductService.create`, que ya es la
unica verdad sobre "cuando un producto es valido" (barcode obligatorio, precio
mayor a cero, sin duplicados). La importacion no reimplementa esas reglas: las
reusa fila por fila y arma un reporte de que entro y que se rechazo y por que.
"""

import csv
import io
import re
import unicodedata

from sqlalchemy.orm import Session

from Backend.services.product_service import ProductService


class ImportError400(Exception):
    """Error que se traduce a HTTP 400: archivo invalido, vacio o no soportado."""


def _norm(texto) -> str:
    """Minusculas, sin acentos y sin espacios sobrantes. Para comparar
    encabezados sin depender de como los escribio cada sistema."""
    s = unicodedata.normalize("NFKD", str(texto or "").strip().lower())
    return "".join(c for c in s if not unicodedata.combining(c))


# Alias de encabezado -> campo interno. El primer encabezado que matchee gana, y
# se consume para que no lo reuse otro campo. El orden de RESOLUCION (abajo)
# resuelve primero los campos menos ambiguos.
ALIASES = {
    "barcode": {
        "barcode", "codigo de barras", "codigo barras", "cod barras",
        "cod. barras", "codigo_barras", "codigo", "cod", "ean", "ean13",
        "upc", "plu", "sku",
    },
    "price": {
        "precio", "price", "pvp", "precio venta", "precio de venta",
        "precio unitario", "importe", "valor", "precio publico",
    },
    "name": {
        "nombre", "name", "producto", "articulo", "descripcion", "detalle",
        "descripcion producto", "descripcion del producto",
    },
    "description": {
        "descripcion larga", "observaciones", "observacion", "notas", "nota",
        "comentario", "comentarios", "detalle adicional",
    },
    "proveedor": {
        "proveedor", "provider", "marca", "fabricante", "supplier",
    },
}

# Orden de resolucion: barcode y precio son inequivocos; el nombre puede caer en
# "descripcion" (Maxirest/Tango exportan el nombre bajo "Descripcion"), asi que
# se resuelve DESPUES, y solo toma "descripcion" si no hubo un "nombre"/"producto"
# explicito.
_RESOLVE_ORDER = ["barcode", "price", "proveedor", "description", "name"]


def _mapear_columnas(encabezados: list) -> dict:
    """Devuelve {campo_interno: indice_de_columna} resolviendo por alias."""
    norm = [_norm(h) for h in encabezados]
    usados = set()
    mapa = {}
    for campo in _RESOLVE_ORDER:
        for i, h in enumerate(norm):
            if i in usados or not h:
                continue
            if h in ALIASES[campo]:
                mapa[campo] = i
                usados.add(i)
                break
    # Red de seguridad: si no se encontro "name" pero quedo una columna
    # "descripcion" libre, esa es el nombre (caso Maxirest/Tango).
    if "name" not in mapa:
        for i, h in enumerate(norm):
            if i not in usados and h in {"descripcion", "detalle"}:
                mapa["name"] = i
                usados.add(i)
                break
    return mapa


def _parsear_precio(valor) -> float:
    """Convierte a float tolerando formato argentino ('1.234,56'), formato
    internacional ('1234.56'), simbolos de moneda y espacios."""
    if valor is None:
        raise ValueError("precio vacio")
    if isinstance(valor, (int, float)):
        return float(valor)

    s = str(valor).strip()
    s = re.sub(r"[^\d,.\-]", "", s)  # saca '$', espacios, etc.
    if s == "":
        raise ValueError("precio vacio")

    # Si tiene coma Y punto, el ultimo separador es el decimal.
    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")   # 1.234,56 -> 1234.56
        else:
            s = s.replace(",", "")                       # 1,234.56 -> 1234.56
    elif "," in s:
        # Solo coma: es el decimal argentino.
        s = s.replace(",", ".")
    return float(s)


def _filas_desde_excel(contenido: bytes) -> list:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise ImportError400(
            "La importacion de Excel no esta disponible (falta openpyxl). "
            "Guardá el archivo como CSV e intentá de nuevo."
        ) from exc
    try:
        wb = load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)
    except Exception as exc:
        raise ImportError400("El archivo Excel es inválido o está dañado.") from exc
    ws = wb.active
    filas = [[c for c in fila] for fila in ws.iter_rows(values_only=True)]
    wb.close()
    return filas


def _filas_desde_csv(contenido: bytes) -> list:
    # Muchos exports de POS argentinos vienen en latin-1 y con ';'.
    for encoding in ("utf-8-sig", "latin-1"):
        try:
            texto = contenido.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ImportError400("No se pudo leer el archivo CSV (codificación desconocida).")

    muestra = texto[:2048]
    try:
        dialecto = csv.Sniffer().sniff(muestra, delimiters=",;\t|")
        delim = dialecto.delimiter
    except csv.Error:
        delim = ";" if muestra.count(";") > muestra.count(",") else ","

    lector = csv.reader(io.StringIO(texto), delimiter=delim)
    return [fila for fila in lector]


def _detectar_y_parsear(filename: str, contenido: bytes) -> list:
    """Devuelve la lista de filas (incluido el encabezado) segun el tipo."""
    nombre = (filename or "").lower()
    if nombre.endswith((".xlsx", ".xlsm", ".xls")):
        return _filas_desde_excel(contenido)
    if nombre.endswith((".csv", ".txt", ".tsv")):
        return _filas_desde_csv(contenido)
    # Sin extension reconocible: probamos Excel (empieza con 'PK') y si no, CSV.
    if contenido[:2] == b"PK":
        return _filas_desde_excel(contenido)
    return _filas_desde_csv(contenido)


def importar_productos(db: Session, filename: str, contenido: bytes) -> dict:
    """Importa productos desde el archivo. Devuelve un resumen:

        {
          "importados": int,
          "rechazados": int,
          "errores": [{"fila": int, "motivo": str}]
        }

    Levanta ImportError400 (que el controlador traduce a HTTP 400) cuando el
    archivo entero es invalido, esta vacio o no tiene columnas reconocibles.
    """
    if not contenido:
        raise ImportError400("El archivo está vacío.")

    filas = _detectar_y_parsear(filename, contenido)
    # Descarto filas totalmente vacias (Excel suele traer sobrantes).
    filas = [f for f in filas if any((c is not None and str(c).strip() != "") for c in f)]

    if len(filas) < 2:
        raise ImportError400("El archivo no contiene registros válidos.")

    encabezados, cuerpo = filas[0], filas[1:]
    mapa = _mapear_columnas(encabezados)

    if "name" not in mapa or "price" not in mapa:
        raise ImportError400(
            "No se reconocieron las columnas mínimas (nombre y precio). "
            "El archivo debe tener una fila de encabezados."
        )

    service = ProductService(db)
    importados = 0
    errores = []

    def _celda(fila, campo):
        idx = mapa.get(campo)
        if idx is None or idx >= len(fila):
            return None
        val = fila[idx]
        if val is None:
            return None
        s = str(val).strip()
        return s if s != "" else None

    for offset, fila in enumerate(cuerpo):
        # +2: la fila 1 es el encabezado y el usuario cuenta desde 1.
        nro = offset + 2
        try:
            nombre = _celda(fila, "name")
            precio_raw = _celda(fila, "price")
            barcode = _celda(fila, "barcode")
            descripcion = _celda(fila, "description")
            proveedor = _celda(fila, "proveedor")

            if not nombre:
                raise ValueError("falta el nombre")
            # El codigo de barras es obligatorio (invariante del dominio de
            # productos): un producto sin codigo no se puede escanear, que es
            # justamente lo que hace esta app. La fila se rechaza con motivo, no
            # se inventa un codigo que el lector nunca va a leer.
            if not barcode:
                raise ValueError("falta el codigo de barras")
            precio = _parsear_precio(precio_raw)

            service.create(barcode, nombre.lower(), precio,
                           descripcion.lower() if descripcion else None,
                           proveedor.lower() if proveedor else None)
            importados += 1
        except ValueError as exc:
            errores.append({"fila": nro, "motivo": str(exc)})
        except Exception as exc:  # fila rota no debe abortar la importacion
            errores.append({"fila": nro, "motivo": f"error inesperado: {exc}"})

    return {
        "importados": importados,
        "rechazados": len(errores),
        "errores": errores,
    }
